"""
paper_fetcher.py — Validate bib entries against Crossref and download PDFs from arXiv.

Provides:
    - query_crossref() to look up a title/author pair via Crossref API
    - try_arxiv_download() to grab a PDF from arXiv when available
    - fetch_all() to batch-process bib entries and produce fetch_report.xlsx
"""

import re
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from thefuzz import fuzz


# ---------------------------------------------------------------------------
# Crossref lookup
# ---------------------------------------------------------------------------

def query_crossref(title: str, authors: str, email: str) -> dict | None:
    """Query Crossref for a work matching *title* and *authors*.

    Parameters
    ----------
    title : str
        The title to search for.
    authors : str
        Author string (used only to refine ranking; Crossref primarily
        matches on title).
    email : str
        Contact email — required by Crossref's polite pool.

    Returns
    -------
    dict | None
        The best-matching item dict from the Crossref response, or None if
        the request fails or returns no results.
    """
    url = "https://api.crossref.org/works"
    params = {
        "query.title": title,
        "rows": 3,
        "mailto": email,
    }
    headers = {
        "User-Agent": f"CitationVerifier/1.0 (mailto:{email})",
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=30)
        resp.raise_for_status()
        items = resp.json().get("message", {}).get("items", [])
        if items:
            return items[0]
        return None
    except Exception as e:
        print(f"  Crossref error for '{title[:60]}…': {e}")
        return None


# ---------------------------------------------------------------------------
# arXiv PDF download
# ---------------------------------------------------------------------------

_ARXIV_DOI_RE = re.compile(r"10\.48550/arXiv\.(\d{4}\.\d{4,5}(?:v\d+)?)")
_ARXIV_URL_RE = re.compile(r"arxiv\.org/(?:abs|pdf)/(\d{4}\.\d{4,5}(?:v\d+)?)")


def _extract_arxiv_id(crossref_result: dict) -> str | None:
    """Try to pull an arXiv ID from a Crossref result dict."""
    # 1. DOI pattern  10.48550/arXiv.XXXX.XXXXX
    doi = crossref_result.get("DOI", "")
    m = _ARXIV_DOI_RE.search(doi)
    if m:
        return m.group(1)

    # 2. Link field — list of {"URL": ...} dicts
    for link in crossref_result.get("link", []):
        url = link.get("URL", "")
        m = _ARXIV_URL_RE.search(url)
        if m:
            return m.group(1)

    # 3. Also check the top-level "URL" and "resource.primary.URL"
    for key in ("URL", ):
        url = crossref_result.get(key, "")
        m = _ARXIV_URL_RE.search(url)
        if m:
            return m.group(1)

    resource_url = (crossref_result.get("resource", {})
                    .get("primary", {})
                    .get("URL", ""))
    m = _ARXIV_URL_RE.search(resource_url)
    if m:
        return m.group(1)

    return None


def _sanitize_filename(title: str, year: str | None) -> str:
    """Build a safe filename from *title* (max 50 chars) + *year*."""
    clean = re.sub(r"[^\w\s-]", "", title.lower())
    clean = re.sub(r"\s+", "_", clean).strip("_")[:50]
    if year:
        clean = f"{clean}_{year}"
    return f"{clean}.pdf"


def try_arxiv_download(crossref_result: dict, papers_dir: str) -> str | None:
    """Download the arXiv PDF for *crossref_result* if an arXiv ID is found.

    Parameters
    ----------
    crossref_result : dict
        A single Crossref item dict.
    papers_dir : str
        Directory to save the PDF.

    Returns
    -------
    str | None
        The saved filename, or None if no arXiv source was found / download
        failed.
    """
    arxiv_id = _extract_arxiv_id(crossref_result)
    if not arxiv_id:
        return None

    # Build filename from crossref title + year
    titles = crossref_result.get("title", [])
    title = titles[0] if titles else arxiv_id
    year = None
    issued = crossref_result.get("issued", {}).get("date-parts", [[]])
    if issued and issued[0]:
        year = str(issued[0][0])

    filename = _sanitize_filename(title, year)
    dest = Path(papers_dir) / filename

    if dest.exists():
        print(f"  Already on disk: {filename}")
        return filename

    pdf_url = f"https://arxiv.org/pdf/{arxiv_id}.pdf"
    try:
        resp = requests.get(pdf_url, timeout=60, stream=True)
        resp.raise_for_status()
        Path(papers_dir).mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        print(f"  Downloaded: {filename}")
        return filename
    except Exception as e:
        print(f"  arXiv download failed ({arxiv_id}): {e}")
        return None


# ---------------------------------------------------------------------------
# Batch processing
# ---------------------------------------------------------------------------

REPORT_COLUMNS = [
    "cite_key",
    "bib_title",
    "crossref_status",
    "crossref_title",
    "doi",
    "pdf_status",
    "pdf_filename",
    "notes",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(bold=True)


def _crossref_title(cr: dict) -> str:
    """Extract the first title string from a Crossref item."""
    titles = cr.get("title", [])
    return titles[0] if titles else ""


def _crossref_doi(cr: dict) -> str:
    return cr.get("DOI", "")


def _write_report(rows: list[dict], output_path: str, summary: dict) -> None:
    """Write fetch_report.xlsx with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fetch Report"

    # Header
    for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        # Conditional fill for the entire row
        status = row.get("crossref_status", "")
        pdf = row.get("pdf_status", "")
        if status == "validated" and pdf == "downloaded":
            fill = GREEN_FILL
        elif status in ("not_found", "low_match") or pdf == "failed":
            fill = RED_FILL
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Summary row
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = BOLD
    summary_items = [
        ("Total", summary["total"]),
        ("Validated", summary["validated"]),
        ("Not found", summary["not_found"]),
        ("PDFs downloaded", summary["downloaded"]),
    ]
    for i, (label, value) in enumerate(summary_items):
        ws.cell(row=summary_row + 1 + i, column=1, value=label)
        ws.cell(row=summary_row + 1 + i, column=2, value=value)

    # Auto column widths
    for col_idx in range(1, len(REPORT_COLUMNS) + 1):
        max_len = len(REPORT_COLUMNS[col_idx - 1])
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


def fetch_all(
    bib_entries: dict,
    papers_dir: str,
    email: str,
) -> None:
    """Validate bib entries against Crossref and download arXiv PDFs.

    Parameters
    ----------
    bib_entries : dict
        ``{cite_key: entry}`` where each entry has a ``.title`` attribute
        (and optionally ``.author``).
    papers_dir : str
        Directory for downloaded PDFs.
    email : str
        Contact email for Crossref polite pool.
    """
    rows: list[dict] = []
    summary = {"total": 0, "validated": 0, "not_found": 0, "downloaded": 0}

    for cite_key, entry in bib_entries.items():
        summary["total"] += 1
        bib_title = entry.title if hasattr(entry, "title") else str(entry)
        authors = entry.authors if hasattr(entry, "authors") else ""

        print(f"[{summary['total']}] {cite_key}: {bib_title[:70]}…")

        row: dict = {
            "cite_key": cite_key,
            "bib_title": bib_title,
            "crossref_status": "",
            "crossref_title": "",
            "doi": "",
            "pdf_status": "skipped",
            "pdf_filename": "",
            "notes": "",
        }

        # --- Crossref lookup ---
        cr = query_crossref(bib_title, authors, email)
        time.sleep(0.1)  # polite delay

        if cr is None:
            row["crossref_status"] = "not_found"
            row["notes"] = "No Crossref result"
            summary["not_found"] += 1
            rows.append(row)
            continue

        cr_title = _crossref_title(cr)
        row["crossref_title"] = cr_title
        row["doi"] = _crossref_doi(cr)

        # --- Title comparison ---
        score = fuzz.token_sort_ratio(bib_title.lower(), cr_title.lower())
        if score >= 80:
            row["crossref_status"] = "validated"
            row["notes"] = f"title_match={score}"
            summary["validated"] += 1
        else:
            row["crossref_status"] = "low_match"
            row["notes"] = f"title_match={score}"
            summary["not_found"] += 1
            rows.append(row)
            continue

        # --- arXiv download ---
        filename = try_arxiv_download(cr, papers_dir)
        if filename:
            row["pdf_status"] = "downloaded"
            row["pdf_filename"] = filename
            summary["downloaded"] += 1
        else:
            row["pdf_status"] = "no_arxiv"

        rows.append(row)

    # --- Write report ---
    report_path = str(Path(papers_dir).parent / "fetch_report.xlsx")
    _write_report(rows, report_path, summary)

    # --- Terminal summary ---
    print("\n" + "=" * 50)
    print("FETCH SUMMARY")
    print("=" * 50)
    print(f"  Total entries:     {summary['total']}")
    print(f"  Crossref validated:{summary['validated']}")
    print(f"  Not found/low:     {summary['not_found']}")
    print(f"  PDFs downloaded:   {summary['downloaded']}")
    print("=" * 50)
