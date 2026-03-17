"""
pdf_mapper.py – Map bib entries to PDF files via fuzzy title matching.

Two-phase workflow:
    1. generate_mapping() – extract titles from PDFs, fuzzy-match to bib entries,
       write mapping.xlsx, then STOP the pipeline for human review.
    2. load_mapping() – read the (possibly hand-edited) mapping.xlsx and return
       a {cite_key: pdf_filename} dict for downstream use.
"""

import asyncio
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from thefuzz import fuzz

from agents import _call_llm


# ---------------------------------------------------------------------------
# Phase 1 – Generate mapping
# ---------------------------------------------------------------------------

async def _extract_title(pdf_path: Path, config: dict, sem: asyncio.Semaphore) -> tuple[str, str]:
    """Extract the title from a PDF's first page using an LLM.

    Returns (pdf_filename, extracted_title).
    """
    async with sem:
        try:
            doc = fitz.open(str(pdf_path))
            page1_text = doc[0].get_text()[:2000] if len(doc) > 0 else ""
            doc.close()
        except Exception:
            return (pdf_path.name, "")

        if not page1_text.strip():
            return (pdf_path.name, "")

        prompt = (
            "Extract ONLY the paper title from this academic paper's first page. "
            "Return just the title, nothing else. No quotes, no explanation.\n\n"
            f"{page1_text}"
        )

        try:
            title = await _call_llm(
                system="You are a helpful assistant.",
                user=prompt,
                config={**config, "max_tokens": 200},
                light=True,
            )
            return (pdf_path.name, title.strip())
        except Exception:
            return (pdf_path.name, "")


async def generate_mapping(
    bib_entries: dict,
    papers_dir: str,
    output_path: str,
    config: dict,
) -> None:
    """Scan PDFs, extract titles, fuzzy-match to bib entries, write mapping.xlsx.

    Parameters
    ----------
    bib_entries : dict
        {cite_key: bib_entry} where each bib_entry has a .title attribute.
    papers_dir : str
        Directory containing the cited PDFs.
    output_path : str
        Path for the output mapping.xlsx file.
    config : dict
        LLM configuration dict (passed to _call_llm).
    """
    papers = Path(papers_dir)
    pdf_files = sorted(papers.glob("*.pdf"))

    if not pdf_files:
        print(f"No PDFs found in {papers_dir}")
        return

    print(f"Found {len(pdf_files)} PDFs. Extracting titles...")

    # Extract titles in parallel with concurrency limit
    sem = asyncio.Semaphore(20)
    tasks = [_extract_title(p, config, sem) for p in pdf_files]
    results = await asyncio.gather(*tasks)

    # Build lookup: {extracted_title: pdf_filename}
    title_to_pdf: dict[str, str] = {}
    for pdf_name, title in results:
        if title:
            title_to_pdf[title] = pdf_name

    print(f"Extracted titles from {len(title_to_pdf)} PDFs.")

    # Fuzzy-match each bib entry to the best PDF title
    rows: list[dict] = []
    for cite_key, entry in bib_entries.items():
        bib_title = entry.title if hasattr(entry, "title") else str(entry)
        best_score = 0
        best_pdf = "??"

        for extracted_title, pdf_name in title_to_pdf.items():
            score = fuzz.token_sort_ratio(bib_title, extracted_title)
            if score > best_score:
                best_score = score
                best_pdf = pdf_name

        status = "matched" if best_score >= 75 else "unmatched"
        rows.append({
            "cite_key": cite_key,
            "bib_title": bib_title,
            "matched_pdf": best_pdf if status == "matched" else "??",
            "match_score": best_score,
            "status": status,
        })

    # Write mapping.xlsx
    _write_mapping_xlsx(rows, output_path)

    # Summary
    matched = sum(1 for r in rows if r["status"] == "matched")
    unmatched = len(rows) - matched
    print(f"\nMapping written to {output_path}")
    print(f"  Matched:   {matched}")
    print(f"  Unmatched: {unmatched}")
    print("\nReview the mapping file, fix any '??' entries, then re-run the pipeline.")


def _write_mapping_xlsx(rows: list[dict], output_path: str) -> None:
    """Write mapping rows to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Mapping"

    columns = ["cite_key", "bib_title", "matched_pdf", "match_score", "status"]
    headers = ["cite_key", "bib_title", "matched_pdf", "match_score", "status"]

    # Write header
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Freeze top row
    ws.freeze_panes = "A2"

    # Conditional fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])

        # Apply conditional formatting per row
        fill = green_fill if row_data["status"] == "matched" else red_fill
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-column-width
    for col_idx in range(1, len(columns) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Phase 2 – Load mapping
# ---------------------------------------------------------------------------

def load_mapping(mapping_path: str) -> dict[str, str]:
    """Read mapping.xlsx and return {cite_key: pdf_filename}.

    Skips rows where matched_pdf is '??' or empty.
    """
    wb = openpyxl.load_workbook(mapping_path, read_only=True)
    ws = wb.active

    mapping: dict[str, str] = {}
    rows = ws.iter_rows(min_row=2, values_only=True)

    for row in rows:
        if row is None or len(row) < 3:
            continue
        cite_key = row[0]
        matched_pdf = row[2]

        if not cite_key or not matched_pdf:
            continue
        if str(matched_pdf).strip() in ("??", ""):
            continue

        mapping[str(cite_key).strip()] = str(matched_pdf).strip()

    wb.close()
    return mapping
